import subprocess, shutil
import os
import re
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
import pyperclip
import asyncio
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl import Workbook
import requests, sys, glob, openpyxl, random
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains


def sanitize_title(title):
    """ Remove special characters from titles to make them safe for file names. """
    # Remove invalid file name characters and strip leading/trailing whitespace
    sanitized = re.sub(r'[\\/*?:"<>|]', '', title).strip()
    return sanitized

def load_data_from_excel(file_path, sheet_name):
    # Load the workbook and select the specified sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Create a list to hold URL and title pairs
    data = []

    # Iterate through the rows in the specified sheet
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=2, values_only=True):
        print("Row data:", row)  # This will show the raw data being read from each row
        url, title = row
        if url and title:
            sanitized_title = sanitize_title(title)
            data.append((url, sanitized_title))
        else:
            print("Missing data in row:", row)  # Indicates which row is missing data

    return data

# Specify the path to your Excel file and the sheet name

file_path = str(input('Nhập tên file:')) + str('.xlsx')
sheet_name = 'Sheet1'
data = load_data_from_excel(file_path, sheet_name)
###
chromium_binary_path = 'C:\\Program Files\\Chromium\\Application\\chrome.exe'
chromium_driver_path = 'C:\\chromedriver\\chromedriver.exe'
service = Service(executable_path=chromium_driver_path)
options = webdriver.ChromeOptions()
options.binary_location = chromium_binary_path
options.add_argument('--disable-features=AutoUpdate')
options.add_argument('--no-sandbox')
options.add_argument('--disable-gpu')
options.add_argument("--ignore-certificate-errors")
options.add_argument("--allow-running-insecure-content")
options.add_experimental_option('excludeSwitches', ['enable-logging'])
# options.add_argument('--headless')

user_name = os.getlogin()
user_data_dir = f'C:\\Users\\{user_name}\\AppData\\Local\\Chromium\\User Data\\'
options.add_argument(f'--user-data-dir={user_data_dir}')
driver = webdriver.Chrome(service=service, options=options)
file_excel = glob.glob('*.xlsx')
# for file in file_excel:
name_folder = file_path.split('.xlsx')[0]
data_folder = '1.Data_file'
text_folder = '1.Text_file_Saved'
audio_folder = '2.Audio'
video_folder = '3.Video_Output'
temp_folder = '4.Temp'
data_folder_path = os.path.join(name_folder,data_folder)
text_folder_path = os.path.join(name_folder,text_folder)
audio_folder_path = os.path.join(name_folder,audio_folder)
video_folder_path = os.path.join(name_folder,video_folder)
temp_folder_path = os.path.join(name_folder,temp_folder)

os.makedirs(name_folder, exist_ok=True)
os.makedirs(data_folder_path, exist_ok=True)
os.makedirs(text_folder_path, exist_ok=True)
os.makedirs(audio_folder_path, exist_ok=True)
os.makedirs(video_folder_path, exist_ok=True)
os.makedirs(temp_folder_path, exist_ok=True)

for url, title in data:
    name_file = title
    output_path_srt = f'{name_folder}/SRT_{name_file}.txt'
    output_path_file_global = f'{name_folder}/{data_folder}/{name_file}.txt'                
    driver.get('https://script.tokaudit.io/')
    insert_link_input = WebDriverWait(driver, 500).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[type="text"][placeholder="Enter Video Url"]')))
    insert_link_input.click()                
    insert_link_input.send_keys(url) #Truyền URL
    start_button = driver.find_element(By.XPATH, '//button[text()="START"]')
    start_button.click() #Click nút Strart
    time.sleep(1)

    try:
        # Chờ đợi cho đến khi thẻ chứa thông báo lỗi xuất hiện
        error_message = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, '//div[contains(text(), "Request failed with status code 404")]')))

        # Nếu thấy thông báo lỗi, thực hiện một yêu cầu khác
        print("Đã có lỗi xảy ra, đang thử lại quá trình")
        # time.sleep(1)
        start_button.click() #Click nút Strart

    except Exception as e:
        # Xử lý ngoại lệ nếu có
        pass


    copy_button = WebDriverWait(driver, 500).until(
    EC.visibility_of_element_located((By.CSS_SELECTOR, "button[class*='bg-pink-600'][class*='text-white'][class*='uppercase']"))
)

    if copy_button == True:
        print('Đã có dữ liệu')
    time.sleep(1)
    # # driver.execute_script("arguments[0].click();", copy_button)
    # time.sleep(2)
    # # x = 1631
    # # y = 704 # Máy ở nhà
    # x = 3182
    # y = 670
    # # pyautogui.click(x, y)
    # # pyautogui.click(x, y)
    # # copied_content = pyperclip.paste()                
    # # with open(output_path_srt, "w") as file:
    # #     file.write(copied_content)
    lines = driver.find_elements(By.CSS_SELECTOR, '.line-content span.text')
    replacement_dictionary = {'No.': 'Number'}
    with open(output_path_file_global, 'w', encoding='utf-8') as file:
        for line in lines:
            print(line.text)
            modified_line = line.text
            for original, replacement in replacement_dictionary.items():
                modified_line = modified_line.replace(original, replacement)
            file.write(modified_line + '\n')
        print(f'Đã xử lý xong ======>>>>>> {name_file}')
shutil.move(file_path, data_folder_path)
print(f'Đã di chuyển file tới ======>>>>>> {data_folder_path}')
driver.quit()