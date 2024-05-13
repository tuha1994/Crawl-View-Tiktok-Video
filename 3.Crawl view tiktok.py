
import subprocess
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
import asyncio
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import random
from openpyxl import Workbook
import requests, sys, re

# Đường dẫn đến trình duyệt Chromium
chromium_binary_path = 'C:\\Program Files\\Chromium\\Application\\chrome.exe'  # Đường dẫn tới trình duyệt Chromium

# Khởi tạo dịch vụ của Chromium WebDriver
chromium_driver_path = 'C:\\chromedriver\\chromedriver.exe'  # Đường dẫn tới Chromium WebDriver
service = Service(executable_path=chromium_driver_path)

# Khởi tạo tùy chọn cho trình duyệt Chromium
options = webdriver.ChromeOptions()
options.binary_location = chromium_binary_path
options.add_argument('--disable-features=AutoUpdate')
options.add_argument('--no-sandbox')
options.add_argument('--disable-gpu')
options.add_argument("--ignore-certificate-errors")
options.add_argument("--allow-running-insecure-content")
options.add_experimental_option('excludeSwitches', ['enable-logging'])
# options.add_experimental_option('useAutomationExtension', False)
# options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_argument('--disable-blink-features=AutomationControlled')



# URL của tệp trên GitHub
github_file_url = 'https://raw.githubusercontent.com/tuha1994/X/main/checkcode.txt'  # Thay thế bằng URL thực tế

# Tải dữ liệu từ GitHub
response = requests.get(github_file_url)

# Hàm làm sạch chuỗi
def clean_string(input_str):
    return input_str.strip().replace(' ', '')  # Loại bỏ khoảng trắng và các ký tự đặc biệt

# Kiểm tra xem yêu cầu tải dữ liệu có thành công không (status code 200 là thành công)
if response.status_code == 200:
    github_data = clean_string(response.text)  # Dữ liệu từ tệp GitHub đã được làm sạch

    # So sánh dữ liệu
    your_data = '2db4336bb7752e79368f490d9048f5ebddassd'  # Thay thế bằng giá trị thực tế
    your_data_cleaned = clean_string(your_data)  # Giá trị của bạn đã được làm sạch
    
    if github_data == your_data_cleaned:
        pass
        # Thực hiện các hành động tiếp theo ở đây
    else:
        print("Error: Server Error >> The server encountered a temporary error and could not complete your request.")
        sys.exit()
        # Bỏ qua quá trình sau hoặc thực hiện các hành động khác nếu cần
else:
    print("Error: Server Error >> The server encountered a temporary error and could not complete your request.")
    sys.exit()
    # Xử lý tình huống không thể tải dữ liệu
user_name = os.getlogin()
#options.add_argument('--headless')
# Chỉ định đường dẫn đến thư mục profile
user_data_dir = f'C:\\Users\\{user_name}\\AppData\\Local\\Chromium\\User Data\\'
options.add_argument(f'--user-data-dir={user_data_dir}')

username = input("Nhập username: ")
driver = webdriver.Chrome(service=service, options=options)
driver.get(f"https://www.tiktok.com/@{username}")
# Chờ cho trang được tải
yy = """

% ·····································································································
% : __    __       ___      .__   __.         ___      .__   __.  __    __     .___________. __    __ :
% :|  |  |  |     /   \     |  \ |  |        /   \     |  \ |  | |  |  |  |    |           ||  |  |  |:
% :|  |__|  |    /  ^  \    |   \|  |       /  ^  \    |   \|  | |  |__|  |    `---|  |----`|  |  |  |:
% :|   __   |   /  /_\  \   |  . `  |      /  /_\  \   |  . `  | |   __   |        |  |     |  |  |  |:
% :|  |  |  |  /  _____  \  |  |\   |     /  _____  \  |  |\   | |  |  |  |        |  |     |  `--'  |:
% :|__|  |__| /__/     \__\ |__| \__|    /__/     \__\ |__| \__| |__|  |__|        |__|      \______/ :
% :                                                                                                   :
% :  ___     ___    ____     ___     ___     ___    ___    _  _       __    _____                     :
% : / _ \   / _ \  |___ \   / _ \   / _ \   / _ \  |__ \  | || |     / /   | ____|                    :
% :| | | | | (_) |   __) | | (_) | | (_) | | (_) |    ) | | || |_   / /_   | |__                      :
% :| | | |  \__, |  |__ <   > _ <   > _ <   > _ <    / /  |__   _| | '_ \  |___ \                     :
% :| |_| |    / /   ___) | | (_) | | (_) | | (_) |  / /_     | |   | (_) |  ___) |                    :
% : \___/    /_/   |____/   \___/   \___/   \___/  |____|    |_|    \___/  |____/                     :
% ·····································································································

"""
print(f"\033[92m{yy}\033[0m")
time.sleep(10)  # Tăng giá trị này nếu cần
# driver.get(f"https://www.tiktok.com/@{username}")
# for i in range(1,30):
#     print(f'Lặp lại lần {i}')
#     # Cuộn đến cuối trang nhiều lần cho đến khi không còn nội dung mới
#     last_height = driver.execute_script("return document.body.scrollHeight")
#     while True:
#         # Cuộn đến cuối trang
#         driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
#         # Chờ cho trang tải nội dung
#         time.sleep(1)
#         # Tính toán chiều cao mới và so sánh với chiều cao trước đó
#         new_height = driver.execute_script("return document.body.scrollHeight")
#         time.sleep(2)
#         if new_height == last_height:
#             break
#         last_height = new_height
def smooth_scroll(driver, start_x, start_y, end_x, end_y, duration=1000):
    script = """
    var startX = arguments[0],
        startY = arguments[1],
        endX = arguments[2],
        endY = arguments[3],
        duration = arguments[4];
    var distanceX = endX - startX,
        distanceY = endY - startY;
    var startTime = new Date().getTime();

    var timer = setInterval(function() {
        var time = new Date().getTime() - startTime,
            newX = startX + (distanceX * (time / duration)),
            newY = startY + (distanceY * (time / duration));
        window.scrollTo(newX, newY);
        
        if (time >= duration) {
            clearInterval(timer);
            window.scrollTo(endX, endY);
        }
    }, 1000/60);  // 60 FPS
    """
    driver.execute_script(script, start_x, start_y, end_x, end_y, duration)
def get_page_height(driver):
    # Lấy chiều cao của trang web
    return driver.execute_script("return document.body.scrollHeight")

def scroll_page_ads(driver, scroll_down_by, down_repeat=200, scroll_threshold=80):
    current_y_position = 0

    for _ in range(down_repeat):
        current_y_position += scroll_down_by
        smooth_scroll(driver, 0, current_y_position - scroll_down_by, 0, current_y_position)
        time.sleep(1.5)

        # Kiểm tra nếu đã cuộn gần cuối trang thì bắt đầu cuộn lên
        if current_y_position + scroll_threshold >= get_page_height(driver):
            break
        

scroll_page_ads(driver, 400)
def clean_illegal_chars(text):
    # Replace non-printable characters (ASCII 0-31 excluding 9, 10, 13) with an empty string
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]+', '', text)

def process_title_string(title_string):
    # Regular expression to find all hashtags
    hashtags = re.findall(r"#\w+", title_string)

    # Find the index of the first hashtag to split the string
    if hashtags:
        first_hashtag_index = title_string.index(hashtags[0])
    else:
        first_hashtag_index = len(title_string)

    # Extract the main title part by slicing to the first hashtag index
    main_title = title_string[:first_hashtag_index].strip()

    # Remove any unwanted trailing parts after 'created by'
    marker_text = ' created by'
    if marker_text in main_title:
        main_title = main_title[:main_title.index(marker_text)].strip()

    # Clean the main title to remove illegal characters
    main_title = clean_illegal_chars(main_title)

    # Join the hashtags into a single string to store in one cell
    hashtags_str = ' '.join(hashtags)

    return main_title, hashtags_str
# Khởi tạo từ điển data với các khóa và danh sách rỗng
data = {'Video URL': [], 'Views': [], 'Number': [], 'Character': [], 'Title': [], 'Hashtags': []}
# Tìm tất cả các links và số lượng view
elements = driver.find_elements(By.CSS_SELECTOR, f'a[href*="/@{username}/video/"]')
number = ""
character = ""
for element in elements:
    video_url = element.get_attribute('href')
    try:
        view_count = element.find_element(By.CSS_SELECTOR, '.video-count').text
    except:
        view_count = 'not found'
    try:
        # Assume the title is extracted from an 'img' tag's 'alt' attribute
        full_title = element.find_element(By.CSS_SELECTOR, 'img').get_attribute('alt')
        title, hashtags = process_title_string(full_title)
    except:
        title = 'Title not found'
        hashtags = ''
    data['Video URL'].append(video_url)
    data['Views'].append(view_count)
    data['Number'].append(number)
    data['Character'].append(character)
    data['Title'].append(title)
    data['Hashtags'].append(hashtags)
driver.quit()
print('Đang xử lý dữ liệu')
# Tạo DataFrame từ dữ liệu
df = pd.DataFrame(data)

# Loại bỏ các dòng có 'Views' là 'not found'
df_filtered = df.query("Views != 'not found'")

# Lưu DataFrame đã lọc vào tệp Excel
output_file_name = f'{username}.xlsx'

# Kiểm tra xem tệp đã tồn tại trong thư mục hiện tại hay chưa
if os.path.exists(output_file_name):
    print(f"Ready...................")
else:
    # Tệp không tồn tại, tạo một tệp mới
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"  # Đặt tên cho sheet nếu muốn
    # Có thể thêm dữ liệu vào sheet ở đây nếu cần
    # Ví dụ: ws['A1'] = "Hello, Excel!"

    # Lưu tệp mới
    wb.save(output_file_name)
    print(f"Tệp '{output_file_name}' đã được tạo mới.")

# Giả sử 'df_filtered' là DataFrame của bạn sau khi đã lọc dữ liệu
df_filtered.to_excel(output_file_name, index=False, engine='openpyxl')

# Mở tệp Excel vừa tạo và điều chỉnh kích thước cột
workbook = load_workbook(output_file_name)
sheet = workbook.active

# Thiết lập kích thước cho cột 1 (A)
sheet.column_dimensions['A'].width = 72.71
sheet.column_dimensions['B'].width = 11
sheet.column_dimensions['C'].width = 13
sheet.column_dimensions['D'].width = 14
sheet.column_dimensions['E'].width = 95

# Đặt bộ lọc cho cột A và B
sheet.auto_filter.ref = "A1:D1"

# Căn chính giữa nội dung cột B
for cell in sheet['B']:
    cell.alignment = Alignment(horizontal='center')
for cell in sheet['C']:
    cell.alignment = Alignment(horizontal='center')
for cell in sheet['D']:
    cell.alignment = Alignment(horizontal='center')


# Lặp qua các ô trong cột B từ dòng thứ 2 trở đi
rows_to_delete = []
for row in range(2, sheet.max_row + 1):
    if not (sheet[f'B{row}'].value and ('M' in sheet[f'B{row}'].value or 'K' in sheet[f'B{row}'].value)):
        rows_to_delete.append(row)
# Đặt công thức cho C2 là LEFT(B2,LEN(B2)-1)
sheet['C2'].value = '=LEFT(B2,LEN(B2)-1)'

# Đặt công thức cho D2 là RIGHT(B2,1)
sheet['D2'].value = '=RIGHT(B2,1)'    


# Tạo một PatternFill để tô màu
fill_color = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")

# Áp dụng màu cho từng ô từ A1 đến D1
for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
    sheet[cell].fill = fill_color


# Xóa các dòng từ dưới lên để tránh thay đổi chỉ số khi xóa
for row in sorted(rows_to_delete, reverse=True):
    sheet.delete_rows(row, 1)
sheet.title = 'DATA'
index_of_data_sheet = workbook.sheetnames.index('DATA')

# Tạo sheet mới tại vị trí trước 'DATA'
new_sheet = workbook.create_sheet('Sheet1', index=index_of_data_sheet)
workbook.save(output_file_name)
print(f'Dữ liệu đã được lưu vào {output_file_name}')